import openpyxl

path ="E:\companies.xlsx"
#file-> workbook -> sheets -> excels
workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet2"]

rows=sheet.max_row
cols=sheet.max_column
print(rows)
print(cols)


for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(r,c).value, end ='     ')
    print()